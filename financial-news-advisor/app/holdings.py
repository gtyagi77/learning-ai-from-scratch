"""Holdings upload: generic template, Zerodha, Groww, Upstox exports.

Accepts both CSV and .xlsx (Zerodha's Console exports are .xlsx by
default) — both are normalized to the same header-sniffing row logic.
Every parser emits lot rows {symbol, quantity, buy_price, buy_date|None};
unresolved or malformed rows are reported back to the user, never
silently dropped. Broker holdings exports usually lack buy dates (only
tradebooks have them), so those lots are date-unknown and treated as
short-term until the user fills the date in.
"""

import csv
import io
import re
from datetime import date, datetime
from typing import Dict, List, Optional, Tuple

import openpyxl

from . import database, tickers

# .xlsx files are ZIP archives; every one starts with this signature.
_XLSX_MAGIC = b"PK\x03\x04"

MAX_ROWS = 2000

GENERIC_TEMPLATE = (
    "symbol,quantity,buy_price,buy_date\n"
    "RELIANCE.NS,10,1450.50,2025-03-12\n"
    "TCS,5,3600,2024-11-02\n"
    "HAL,12,4100,\n"
)


def _norm_header(h: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (h or "").lower())


def _num(v: Optional[str]) -> Optional[float]:
    if v is None:
        return None
    v = v.replace(",", "").replace("₹", "").strip()
    if not v:
        return None
    try:
        return float(v)
    except ValueError:
        return None


def _date(v: Optional[str]) -> Optional[str]:
    if not v or not v.strip():
        return None
    v = v.strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d-%b-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(v, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _resolve(symbol_or_name: str) -> Optional[str]:
    """Resolve a broker symbol or company name to a Yahoo ticker."""
    raw = (symbol_or_name or "").strip()
    if not raw:
        return None
    up = raw.upper()
    if up.endswith(".NS") or up.endswith(".BO") or re.match(r"^[A-Z0-9&\-]{1,20}$", up):
        resolved = tickers.resolve_symbol(up)
        if resolved.endswith(".NS") or resolved.endswith(".BO"):
            return resolved
        if resolved in tickers.US_BASES:
            return resolved  # known US large-cap (AAPL, BRK-B)
        # Bare Indian-looking symbol not in the known map: assume NSE.
        if re.match(r"^[A-Z0-9&\-]{2,20}$", up):
            return up + ".NS"
    # Company name (Groww exports names): look it up in the alias map.
    low = raw.lower()
    for name, sym in tickers.COMPANY_MAP.items():
        if name == low or (len(name) > 4 and name in low):
            return sym
    return None


# ---------------------------------------------------------------------------
# format detection + per-format row mapping
# ---------------------------------------------------------------------------

FORMATS = {
    # format: (required normalized headers, field mapping)
    "generic":  ({"symbol", "quantity", "buyprice"},
                 {"symbol": "symbol", "quantity": "quantity",
                  "buyprice": "buy_price", "buydate": "buy_date"}),
    "zerodha_tradebook": ({"tradedate", "tradingsymbol", "tradetype", "quantity", "price"}, None),
    "zerodha_holdings": ({"instrument", "qty", "avgcost"},
                         {"instrument": "symbol", "qty": "quantity",
                          "avgcost": "buy_price"}),
    "groww":    ({"stockname", "quantity", "averagebuyprice"},
                 {"stockname": "symbol", "quantity": "quantity",
                  "averagebuyprice": "buy_price"}),
    "upstox":   ({"instrument", "quantity", "avgprice"},
                 {"instrument": "symbol", "quantity": "quantity",
                  "avgprice": "buy_price"}),
}
_HEADER_ALIASES = {
    "instrument": ("instrument", "symbol"),
    "qty": ("qty", "qty.", "quantity", "quantity available"),
    "avgcost": ("avgcost", "avg.cost", "avgcost.", "average price"),
    "avgprice": ("avgprice", "averageprice", "avg.price"),
    "stockname": ("stockname", "companyname", "scripname"),
    "averagebuyprice": ("averagebuyprice", "avgbuyprice", "buyaverageprice"),
}


def _detect_format(headers: List[str]) -> Optional[str]:
    norm = {_norm_header(h) for h in headers}
    # expand aliases into canonical names
    canon = set(norm)
    for canonical, variants in _HEADER_ALIASES.items():
        if any(_norm_header(v) in norm for v in variants):
            canon.add(canonical)
    for fmt, (required, _) in FORMATS.items():
        if required <= canon:
            return fmt
    return None


def _header_value(row: Dict[str, str], *names: str) -> Optional[str]:
    for k, v in row.items():
        if _norm_header(k) in {_norm_header(n) for n in names}:
            return v
    return None


def _cell_to_str(v) -> str:
    """Normalize one openpyxl cell value to the same plain-string shape
    csv.DictReader already hands the rest of this module — so format
    detection, _num, _date etc. work identically regardless of source."""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, float) and v.is_integer():
        return str(int(v))  # avoid "10.0" for share counts
    return str(v)


_HEADER_SCAN_ROWS = 50


def _select_xlsx_sheet(wb):
    """Some Zerodha exports bundle multiple sheets (Equity / Mutual Funds /
    Combined) -- prefer the one literally named "Equity" over whichever
    sheet happens to be marked active, since that's the one with tickers
    this app can price. Falls back to wb.active for single-sheet exports."""
    for name in wb.sheetnames:
        if name.strip().lower() == "equity":
            return wb[name]
    return wb.active


def _rows_from_xlsx(raw: bytes) -> Tuple[List[str], List[Dict[str, str]]]:
    # read_only=True (streaming mode) has been observed to silently
    # truncate some real-world exports to a single row -- uploads are
    # capped at MAX_UPLOAD_BYTES anyway, so loading fully into memory
    # costs nothing meaningful and is the more robust choice here.
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    try:
        sheet = _select_xlsx_sheet(wb)
        all_rows = list(sheet.iter_rows(values_only=True))
        if not all_rows:
            return [], []

        # Broker exports (Zerodha's Console .xlsx included) commonly prepend
        # a title/account-info block before the real header row, so don't
        # assume row 1 is it — scan for the first row _detect_format
        # actually recognizes. Falls back to row 1 if nothing matches, so
        # an unrecognized file still reports real header text.
        header_idx = 0
        for i, row in enumerate(all_rows[:_HEADER_SCAN_ROWS]):
            candidate = [str(h).strip() if h is not None else "" for h in row]
            if _detect_format(candidate):
                header_idx = i
                break

        headers = [str(h).strip() if h is not None else "" for h in all_rows[header_idx]]
        rows: List[Dict[str, str]] = []
        for values in all_rows[header_idx + 1:]:
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in values):
                continue  # skip fully blank rows (common trailing rows in exports)
            row = {h: _cell_to_str(v) for h, v in zip(headers, values) if h}
            rows.append(row)
            if len(rows) >= MAX_ROWS:
                break
        return headers, rows
    finally:
        wb.close()


def _process_rows(headers: List[str], rows: List[Dict[str, str]]) -> Tuple[List[Dict], str, List[str]]:
    errors: List[str] = []
    fmt = _detect_format(headers)
    if not fmt:
        return [], "unknown", [
            "unrecognized columns — expected the generic template "
            "(symbol,quantity,buy_price,buy_date) or a Zerodha / Groww / "
            f"Upstox export. Found: {', '.join(headers[:8])}"]

    if fmt == "zerodha_tradebook":
        return _parse_tradebook(rows)

    lots: List[Dict] = []
    for i, row in enumerate(rows, start=2):
        if fmt == "generic":
            sym_raw = _header_value(row, "symbol")
            qty = _num(_header_value(row, "quantity"))
            price = _num(_header_value(row, "buy_price", "buyprice"))
            bdate = _date(_header_value(row, "buy_date", "buydate"))
        elif fmt == "zerodha_holdings":
            sym_raw = _header_value(row, "instrument", "symbol")
            qty = _num(_header_value(row, "qty", "qty.", "quantity", "quantity available"))
            price = _num(_header_value(row, "avg cost", "avg. cost", "avgcost", "average price"))
            bdate = None
        elif fmt == "groww":
            sym_raw = _header_value(row, "stock name", "company name", "scrip name")
            qty = _num(_header_value(row, "quantity"))
            price = _num(_header_value(row, "average buy price", "avg buy price"))
            bdate = None
        else:  # upstox
            sym_raw = _header_value(row, "instrument", "symbol")
            qty = _num(_header_value(row, "quantity"))
            price = _num(_header_value(row, "avg price", "average price", "avgprice"))
            bdate = None

        if not sym_raw or qty is None or price is None:
            if any((row.get(k) or "").strip() for k in row):
                errors.append(f"row {i}: missing symbol/quantity/price — skipped")
            continue
        if qty <= 0 or price <= 0:
            errors.append(f"row {i}: non-positive quantity/price — skipped")
            continue
        symbol = _resolve(sym_raw)
        if not symbol:
            errors.append(f"row {i}: could not resolve '{sym_raw}' to a ticker — skipped")
            continue
        lots.append({"symbol": symbol, "quantity": qty, "buy_price": price,
                     "buy_date": bdate})
    return lots, fmt, errors


def parse_csv(content: str) -> Tuple[List[Dict], str, List[str]]:
    """(lots, format_name, errors). Lots: symbol/quantity/buy_price/buy_date."""
    try:
        reader = csv.DictReader(io.StringIO(content))
        headers = reader.fieldnames or []
        rows = list(reader)[:MAX_ROWS]
    except csv.Error as exc:
        return [], "unknown", [f"could not parse CSV: {exc}"]
    return _process_rows(headers, rows)


def parse_xlsx(raw: bytes) -> Tuple[List[Dict], str, List[str]]:
    """Same as parse_csv but for .xlsx — Zerodha's Console exports default
    to .xlsx, not CSV."""
    try:
        headers, rows = _rows_from_xlsx(raw)
    except Exception as exc:
        return [], "unknown", [f"could not read spreadsheet: {exc}"]
    return _process_rows(headers, rows)


def parse_holdings_file(raw: bytes) -> Tuple[List[Dict], str, List[str]]:
    """Upload entry point: detects .xlsx (by its ZIP magic bytes) vs CSV
    from the raw upload and parses accordingly."""
    if raw[:4] == _XLSX_MAGIC:
        return parse_xlsx(raw)
    try:
        content = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        return [], "unknown", ["file is not a recognized .csv or .xlsx file"]
    return parse_csv(content)


def _parse_tradebook(rows: List[Dict]) -> Tuple[List[Dict], str, List[str]]:
    """Zerodha tradebook: buys create lots; sells consume them FIFO, so the
    result is the surviving lots with accurate buy dates."""
    errors: List[str] = []
    trades = []
    for i, row in enumerate(rows, start=2):
        sym_raw = _header_value(row, "tradingsymbol", "symbol")
        ttype = (_header_value(row, "trade_type", "tradetype") or "").lower()
        qty = _num(_header_value(row, "quantity", "qty"))
        price = _num(_header_value(row, "price", "avg price"))
        tdate = _date(_header_value(row, "trade_date", "tradedate", "order_execution_time"))
        if not sym_raw or qty is None or price is None or ttype not in ("buy", "sell"):
            if any((row.get(k) or "").strip() for k in row):
                errors.append(f"row {i}: incomplete trade — skipped")
            continue
        symbol = _resolve(sym_raw)
        if not symbol:
            errors.append(f"row {i}: could not resolve '{sym_raw}' — skipped")
            continue
        trades.append({"symbol": symbol, "type": ttype, "quantity": qty,
                       "price": price, "date": tdate or ""})

    trades.sort(key=lambda t: t["date"])
    open_lots: Dict[str, List[Dict]] = {}
    for t in trades:
        lots = open_lots.setdefault(t["symbol"], [])
        if t["type"] == "buy":
            lots.append({"symbol": t["symbol"], "quantity": t["quantity"],
                         "buy_price": t["price"], "buy_date": t["date"] or None})
        else:  # FIFO sell
            remaining = t["quantity"]
            while remaining > 1e-9 and lots:
                lot = lots[0]
                take = min(lot["quantity"], remaining)
                lot["quantity"] -= take
                remaining -= take
                if lot["quantity"] <= 1e-9:
                    lots.pop(0)
            if remaining > 1e-9:
                errors.append(f"{t['symbol']}: sold more than bought in this "
                              f"file — excess {remaining:g} ignored")
    result = [lot for lots in open_lots.values() for lot in lots
              if lot["quantity"] > 1e-9]
    return result, "zerodha_tradebook", errors


def import_lots(user_id: int, lots: List[Dict], source: str,
                replace: bool = True) -> Dict:
    """Store parsed lots for a user; ensure each ticker is in the portfolio
    so it gets news coverage and recommendations."""
    if replace:
        database.clear_lots(user_id, source=source)
    for lot in lots:
        database.add_lot(user_id, lot["symbol"], lot["quantity"],
                         lot["buy_price"], lot.get("buy_date"), source=source)
        database.upsert_holding(lot["symbol"], None, user_id=user_id)
    return {"imported": len(lots)}
