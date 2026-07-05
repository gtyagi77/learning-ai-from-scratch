import os
import sys
from datetime import date, timedelta

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app import config, database, holdings, taxes  # noqa: E402
from tests.conftest import make_authed_client  # noqa: E402


def setup_module():
    database.init(":memory:")


def _days_ago(n: int) -> str:
    return (date.today() - timedelta(days=n)).isoformat()


# ---------------- CSV parsers ----------------

def test_parse_generic_csv():
    body = ("symbol,quantity,buy_price,buy_date\n"
            "RELIANCE.NS,10,1450.50,2025-03-12\n"
            "tcs,5,3600,12-11-2024\n"          # bare symbol + dd-mm-yyyy
            "HAL,12,4100,\n"                    # missing date is fine
            ",3,100,2025-01-01\n")              # missing symbol -> error
    lots, fmt, errors = holdings.parse_csv(body)
    assert fmt == "generic"
    assert len(lots) == 3
    by_sym = {l["symbol"]: l for l in lots}
    assert by_sym["RELIANCE.NS"]["buy_date"] == "2025-03-12"
    assert by_sym["TCS.NS"]["buy_date"] == "2024-11-12"
    assert by_sym["HAL.NS"]["buy_date"] is None
    assert len(errors) == 1 and "row 5" in errors[0]


def test_parse_zerodha_holdings():
    body = ("Instrument,Qty.,Avg. cost,LTP,Cur. val,P&L\n"
            "RELIANCE,10,\"1,450.50\",1528.4,15284,779\n"
            "TATAMOTORS,20,795.00,688.15,13763,-2137\n")
    lots, fmt, errors = holdings.parse_csv(body)
    assert fmt == "zerodha_holdings"
    assert {l["symbol"] for l in lots} == {"RELIANCE.NS", "TATAMOTORS.NS"}
    assert all(l["buy_date"] is None for l in lots)  # holdings export: no dates
    assert not errors


def test_parse_zerodha_tradebook_fifo_netting():
    body = ("trade_date,tradingsymbol,trade_type,quantity,price\n"
            "2024-05-10,INFY,buy,10,1400\n"
            "2024-08-01,INFY,buy,10,1500\n"
            "2025-01-15,INFY,sell,12,1650\n")   # consumes lot1 + 2 of lot2
    lots, fmt, errors = holdings.parse_csv(body)
    assert fmt == "zerodha_tradebook"
    assert len(lots) == 1
    lot = lots[0]
    assert lot["symbol"] == "INFY.NS"
    assert lot["quantity"] == 8
    assert lot["buy_price"] == 1500
    assert lot["buy_date"] == "2024-08-01"
    assert not errors


def test_parse_groww_and_upstox():
    groww = ("Stock Name,ISIN,Quantity,Average Buy Price\n"
             "Reliance Industries,INE002A01018,4,1440\n"
             "Some Unknown Co,INE000X00000,2,100\n")
    lots, fmt, errors = holdings.parse_csv(groww)
    assert fmt == "groww"
    assert len(lots) == 1 and lots[0]["symbol"] == "RELIANCE.NS"
    assert any("Some Unknown Co" in e for e in errors)

    upstox = ("Instrument,Quantity,Avg Price\nHAL,6,3900\n")
    lots2, fmt2, _ = holdings.parse_csv(upstox)
    assert fmt2 == "upstox" and lots2[0]["symbol"] == "HAL.NS"


def test_unknown_format_reports_headers():
    lots, fmt, errors = holdings.parse_csv("foo,bar\n1,2\n")
    assert fmt == "unknown" and not lots
    assert "unrecognized columns" in errors[0]


# ---------------- .xlsx parser (Zerodha Console exports default to .xlsx) ----------------

def _make_xlsx(headers, rows, preamble=None) -> bytes:
    import io as _io

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    for pre_row in (preamble or []):
        ws.append(pre_row)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_xlsx_zerodha_holdings():
    raw = _make_xlsx(
        ["Instrument", "Qty.", "Avg. cost", "LTP", "Cur. val", "P&L"],
        [["RELIANCE", 10, 1450.50, 1528.4, 15284, 779],
         ["TATAMOTORS", 20, 795.00, 688.15, 13763, -2137]],
    )
    lots, fmt, errors = holdings.parse_xlsx(raw)
    assert fmt == "zerodha_holdings"
    assert {l["symbol"] for l in lots} == {"RELIANCE.NS", "TATAMOTORS.NS"}
    assert not errors


def test_parse_xlsx_skips_title_preamble_before_header_row():
    # Real Zerodha Console exports prepend a title/account-info block
    # before the actual column headers -- the header row is not row 1.
    raw = _make_xlsx(
        ["Instrument", "Qty.", "Avg. cost", "LTP", "Cur. val", "P&L"],
        [["RELIANCE", 10, 1450.50, 1528.4, 15284, 779]],
        preamble=[["Holdings statement for XX1234 as on 05-Jul-2026"],
                  [None, None, None]],
    )
    lots, fmt, errors = holdings.parse_xlsx(raw)
    assert fmt == "zerodha_holdings"
    assert lots and lots[0]["symbol"] == "RELIANCE.NS"
    assert not errors


def test_parse_xlsx_zerodha_portfolio_statement_shape():
    # Mirrors the real shape of Zerodha's newer multi-sheet "portfolio
    # statement" export: Equity/Mutual Funds/Combined sheets, a ~22-row
    # preamble (guide text, Client ID, report title, a Summary stats
    # block) before the real header row, and different column names
    # (Symbol/Quantity Available/Average Price) than the classic
    # Instrument/Qty./Avg. cost holdings export. All data below is
    # synthetic, not the real uploaded file.
    import io as _io

    from openpyxl import Workbook

    wb = Workbook()
    equity = wb.active
    equity.title = "Equity"
    preamble = [
        [None], [None], [None],
        [None, "View Zerodha's guide on using tax reports for filing."],
        [None], [None],
        [None, "Client ID", "TEST0000"],
        [None], [None], [None],
        [None, "Equity Holdings Statement as on 2026-01-01"],
        [None],
        [None, "Summary"],
        [None],
        [None, "Invested Value", 100000],
        [None, "Present Value", 110000],
        [None, "Unrealized P&L", 10000],
        [None, "Unrealized P&L Pct.", 10.0],
        [None], [None],
        [None, ""],
        [None],
    ]
    for row in preamble:
        equity.append(row)
    equity.append([None, "Symbol", "ISIN", "Sector", "Quantity Available",
                   "Quantity Discrepant", "Quantity Long Term",
                   "Quantity Pledged (Margin)", "Quantity Pledged (Loan)",
                   "Average Price", "Previous Closing Price",
                   "Unrealized P&L", "Unrealized P&L Pct."])
    equity.append([None, "RELIANCE", "INE002A01018", "ENERGY", 10, 0, 10,
                   0, 0, 1450.5, 1500.0, 500, 3.4])
    equity.append([None, "TCS", "INE467B01029", "IT", 5, 0, 5,
                   0, 0, 3600.0, 3650.0, 250, 1.4])

    mf = wb.create_sheet("Mutual Funds")
    mf.append(["Fund Name", "Units", "NAV"])
    mf.append(["Some Fund", 100, 25.5])

    wb.create_sheet("Combined")

    buf = _io.BytesIO()
    wb.save(buf)

    lots, fmt, errors = holdings.parse_xlsx(buf.getvalue())
    assert fmt == "zerodha_holdings"
    assert not errors
    by_sym = {l["symbol"]: l for l in lots}
    assert set(by_sym) == {"RELIANCE.NS", "TCS.NS"}
    assert by_sym["RELIANCE.NS"]["quantity"] == 10
    assert by_sym["RELIANCE.NS"]["buy_price"] == 1450.5
    assert by_sym["TCS.NS"]["quantity"] == 5
    # The Mutual Funds sheet's "Fund Name"/"Units"/"NAV" columns never
    # leak into the parsed Equity lots.
    assert "Some Fund" not in str(lots)


def test_parse_xlsx_still_reports_unrecognized_columns_with_no_match():
    raw = _make_xlsx(["foo", "bar"], [[1, 2]])
    lots, fmt, errors = holdings.parse_xlsx(raw)
    assert fmt == "unknown" and not lots
    assert "unrecognized columns" in errors[0]
    assert "foo" in errors[0] and "bar" in errors[0]


def test_parse_xlsx_with_native_date_cells():
    import datetime as _dt

    raw = _make_xlsx(
        ["symbol", "quantity", "buy_price", "buy_date"],
        [["RELIANCE.NS", 10, 1450.50, _dt.date(2025, 3, 12)],
         ["HAL", 12, 4100, None]],
    )
    lots, fmt, errors = holdings.parse_xlsx(raw)
    assert fmt == "generic"
    by_sym = {l["symbol"]: l for l in lots}
    assert by_sym["RELIANCE.NS"]["buy_date"] == "2025-03-12"
    assert by_sym["HAL.NS"]["buy_date"] is None
    assert not errors


def test_parse_xlsx_skips_trailing_blank_rows():
    raw = _make_xlsx(
        ["Instrument", "Qty.", "Avg. cost"],
        [["HAL", 6, 3900], [None, None, None], [None, None, None]],
    )
    lots, fmt, errors = holdings.parse_xlsx(raw)
    assert fmt == "zerodha_holdings"
    assert len(lots) == 1 and lots[0]["symbol"] == "HAL.NS"
    assert not errors


def test_parse_holdings_file_dispatches_by_magic_bytes():
    xlsx_raw = _make_xlsx(["Instrument", "Qty.", "Avg. cost"], [["HAL", 6, 3900]])
    lots, fmt, errors = holdings.parse_holdings_file(xlsx_raw)
    assert fmt == "zerodha_holdings" and lots[0]["symbol"] == "HAL.NS"

    csv_raw = b"symbol,quantity,buy_price,buy_date\nRELIANCE.NS,10,1450.50,2025-03-12\n"
    lots2, fmt2, _ = holdings.parse_holdings_file(csv_raw)
    assert fmt2 == "generic" and lots2[0]["symbol"] == "RELIANCE.NS"

    lots3, fmt3, errors3 = holdings.parse_holdings_file(b"\xff\xfe\x00garbage-not-utf8\x80")
    assert fmt3 == "unknown" and not lots3 and errors3


# ---------------- taxes ----------------

def test_st_lt_split_and_tax():
    lots = [
        {"quantity": 10, "buy_price": 100, "buy_date": _days_ago(400)},  # LT
        {"quantity": 10, "buy_price": 120, "buy_date": _days_ago(100)},  # ST
    ]
    pos = taxes.analyze_position("XYZ.NS", lots, current_price=150.0)
    assert pos["lt_quantity"] == 10 and pos["st_quantity"] == 10
    assert pos["lt_gain"] == 500.0   # (150-100)*10
    assert pos["st_gain"] == 300.0   # (150-120)*10
    expected_tax = 300 * config.TAX_STCG_RATE + 500 * config.TAX_LTCG_RATE
    assert pos["tax_if_sold_today"] == round(expected_tax, 2)
    # ST lot turns LT (TAX_LT_DAYS+1) days after purchase.
    assert pos["next_lt_days"] == (config.TAX_LT_DAYS + 1) - 100
    assert pos["tax_saved_by_waiting"] == round(
        300 * (config.TAX_STCG_RATE - config.TAX_LTCG_RATE), 2)


def test_unknown_dates_are_short_term():
    lots = [{"quantity": 5, "buy_price": 100, "buy_date": None}]
    pos = taxes.analyze_position("ABC.NS", lots, current_price=140.0)
    assert pos["st_quantity"] == 5 and pos["date_unknown_quantity"] == 5
    assert pos["tax_if_sold_today"] == round(200 * config.TAX_STCG_RATE, 2)


def test_portfolio_summary_applies_exemption_once():
    lots = [{"quantity": 100, "buy_price": 1000, "buy_date": _days_ago(500)}]
    p1 = taxes.analyze_position("BIGWIN.NS", lots, current_price=3000.0)  # 2L LT gain
    summary = taxes.portfolio_summary([p1])
    assert summary["lt_gain"] == 200000.0
    assert summary["ltcg_exemption_applied"] == config.TAX_LTCG_EXEMPTION
    assert summary["tax_if_all_sold_today"] == round(
        (200000 - config.TAX_LTCG_EXEMPTION) * config.TAX_LTCG_RATE, 2)


def test_portfolio_summary_reports_unknown_tax_when_no_quotes_available():
    # With no live price for any position, gains are computed as 0 for
    # every lot (see analyze_position's `if current_price else 0.0`) --
    # tax/exemption must come back as None ("unknown"), not a confident
    # ₹0.00, matching current_value/unrealized_gain's existing behavior.
    lots = [{"quantity": 100, "buy_price": 1000, "buy_date": _days_ago(500)}]
    p1 = taxes.analyze_position("NOPRICE.NS", lots, current_price=None)
    summary = taxes.portfolio_summary([p1])
    assert summary["current_value"] is None
    assert summary["tax_if_all_sold_today"] is None
    assert summary["ltcg_exemption_applied"] is None


def test_sell_moderated_to_hold_when_waiting_beats_downside():
    # ST lot, 30 days from turning LT, big gain -> tax saved by waiting is
    # large; expected downside is small -> HOLD with a dated note.
    lots = [{"quantity": 100, "buy_price": 100,
             "buy_date": _days_ago(config.TAX_LT_DAYS - 29)}]
    pos = taxes.analyze_position("WAIT.NS", lots, current_price=200.0)
    assert pos["next_lt_days"] == 30
    action, note = taxes.tax_adjust_action("SELL", -1.0, pos)  # 1% downside
    assert action == "HOLD"
    assert note and "saves" in note and pos["next_lt_date"] in note

    # Big expected downside outweighs the tax saving -> SELL stands.
    action2, note2 = taxes.tax_adjust_action("SELL", -9.0, pos)
    assert action2 == "SELL" and note2 and "outweighs" in note2

    # Far from LT (>60 days) -> no moderation.
    lots_far = [{"quantity": 100, "buy_price": 100, "buy_date": _days_ago(10)}]
    pos_far = taxes.analyze_position("FAR.NS", lots_far, current_price=200.0)
    action3, note3 = taxes.tax_adjust_action("SELL", -1.0, pos_far)
    assert action3 == "SELL" and note3 is None


def test_buy_actions_untouched_by_tax():
    lots = [{"quantity": 10, "buy_price": 100, "buy_date": _days_ago(100)}]
    pos = taxes.analyze_position("KEEP.NS", lots, current_price=150.0)
    assert taxes.tax_adjust_action("STRONG BUY", 5.0, pos) == ("STRONG BUY", None)


# ---------------- upload endpoint end-to-end ----------------

def test_upload_and_holdings_view(monkeypatch):
    from app import prices
    monkeypatch.setattr(prices, "get_quote",
                        lambda s: {"price": 2000.0, "previous_close": 1980.0,
                                   "currency": "INR", "change_pct": 1.0})
    client = make_authed_client()
    body = ("symbol,quantity,buy_price,buy_date\n"
            f"RELIANCE.NS,10,1500,{_days_ago(400)}\n"
            f"RELIANCE.NS,5,1800,{_days_ago(60)}\n")
    r = client.post("/api/holdings/upload", content=body,
                    headers={"Content-Type": "text/csv"})
    assert r.status_code == 200 and r.json()["imported"] == 2

    view = client.get("/api/holdings").json()
    pos = view["positions"][0]
    assert pos["ticker"] == "RELIANCE.NS"
    assert pos["quantity"] == 15
    assert pos["lt_quantity"] == 10 and pos["st_quantity"] == 5
    assert view["summary"]["invested"] == 10 * 1500 + 5 * 1800
    assert "not tax advice" in view["summary"]["note"]

    # Re-upload with the same source replaces, not duplicates.
    r2 = client.post("/api/holdings/upload", content=body,
                     headers={"Content-Type": "text/csv"})
    assert r2.status_code == 200
    assert client.get("/api/holdings").json()["positions"][0]["quantity"] == 15

    # Template download works and is CSV.
    t = client.get("/api/holdings/template")
    assert t.status_code == 200 and t.text.startswith("symbol,quantity")
